from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import os

#Path to chrome driver and Xpath to div container that contains top 10 tweets
PATH = r'Drivers\chromedriver.exe'
twitter_XPATH = "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[2]/div/div/section/div/div"
driver = webdriver.Chrome(PATH)
driver.get("https://twitter.com/explore/tabs/trending")

# wait to do any processing until access to XPATH is confirmed
try:
    main = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, twitter_XPATH))
    )
    #create array for data
    content_array = []

    #Write each line to array, not including '·' characters
    for line in main.text.split('\n'):
        if line != '·':
            content_array.append(line)

    del content_array[:2]

    #try to open file if it exists, otherwise create new excel file
    if os.path.exists('twitter_trends.xlsx'):
        wb = openpyxl.load_workbook(r'twitter_trends.xlsx')
    else:
        wb = openpyxl.Workbook()
        wb.save(r'twitter_trends.xlsx')

    #Create new work sheet in open exel file and name it with today's date and time
    now = datetime.now()
    dt_string = now.strftime("%d_%m_%Y__%Hh_%Mm_%Ss")
    wb.create_sheet(dt_string, 0)

    #set sheet as active work sheet
    sheet = wb.active

    #Declare variables to identify excel position
    #index starts at 2 to trim wasted space
    index = 2
    column = "A"
    row = 1

    #Split tweets to separate columns when tweet identifying number is found
    for x in content_array:
        if x != str(index):
            active_cell = column + str(row)
            sheet[active_cell].alignment = Alignment(wrap_text=True)
            sheet.column_dimensions[column].width = 25
            sheet[active_cell] = x
            row += 1
        else:
            #reset row position, increment index and excel letter row by one
            index += 1
            row = 1
            column = chr(ord(column) + 1)

    wb.save(r'twitter_trends.xlsx')
finally:
    driver.quit()
